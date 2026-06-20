"""
Módulo de Valuation semi-automatizado — metodologia Parreira (Zelen Invest).

Modelo forward por múltiplo de saída + TIR:
  1. Projeta Receita (trajetória de crescimento) → EBIT (margem) por N anos.
  2. Valor de saída = EV/EBIT_saída × EBIT_terminal − Dívida Líquida_saída.
  3. Fluxos: entrada = −(Preço×Ações + Dív.Líq) [compra o EV]; intermediários =
     dividendos; saída = valor do equity.
  4. TIR = IRR dos fluxos; Preço Justo = VP da saída ÷ ações; Upside vs preço atual.
  5. Bloco de comparáveis setoriais (P/E, EV/EBIT, ROE, Cresc.) vs média de pares.

"Semi-automatizado": as ÂNCORAS (preço, ações, dívida líquida, receita/EBIT LTM,
EV/EBIT atual) são preenchidas pelo pipeline (CVM + FRE + Fundamentus); as
PREMISSAS (crescimento, margem, múltiplo de saída, desconto, pares) vêm com
defaults sensatos e ficam editáveis — no Excel gerado e nos argumentos.

Saída: Excel no layout do playbook, com fórmulas vivas (recalcula ao abrir).

CLI:
    python valuation.py --ticker WEGE3
    python valuation.py --ticker WEGE3 --anos 5 --saida-mult 12 --desconto 0.15
"""

from __future__ import annotations

import argparse
import math
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional


# ─── Estruturas ───────────────────────────────────────────────────────────────

@dataclass
class Anchors:
    """Dados de mercado/contábeis preenchidos automaticamente (em R$ milhões)."""
    ticker: str
    nome: str = ""
    preco: float = 0.0           # R$/ação
    n_acoes: float = 0.0         # milhões de ações
    net_debt: float = 0.0        # R$ mi (dívida bruta − caixa)
    receita_ltm: float = 0.0     # R$ mi
    ebit_ltm: float = 0.0        # R$ mi
    margem_ebit: float = 0.0     # ebit/receita
    ev_ebit_atual: Optional[float] = None
    roic: Optional[float] = None  # ROIC em FRAÇÃO (0.15 = 15%) — dirige o reinvestimento perpétuo
    # ── Insumos do build-up FCFF (LTM, R$ mi) ──────────────────────────────────
    da_ltm: float = 0.0          # Depreciação & Amortização
    capex_ltm: float = 0.0       # Capex (investimento)
    cogs_ltm: float = 0.0        # Custo dos produtos vendidos (|CMV|) — base de DIO/DPO
    receber: float = 0.0         # Contas a receber (capital de giro)
    estoque: float = 0.0         # Estoques
    fornecedores: float = 0.0    # Fornecedores
    # ── Insumos do FCFE (financeiras: bancos, seguradoras, B3/corretoras) ───────
    lucro_liq_ltm: float = 0.0   # Lucro líquido atribuível aos controladores (R$ mi)
    pl: float = 0.0              # Patrimônio líquido contábil (book equity, R$ mi)
    # Comparáveis (empresa analisada)
    pe_atual: Optional[float] = None
    roe: Optional[float] = None
    cagr_hist: Optional[float] = None
    fontes: Dict[str, str] = field(default_factory=dict)


@dataclass
class Premissas:
    """Premissas do modelo (editáveis). Arrays indexados por ano 0..anos."""
    anos: int = 5               # horizonte explícito (5a — além disso vira especulação)
    base_year: int = field(default_factory=lambda: date.today().year)
    cresc: List[float] = field(default_factory=list)     # [_, g1..gN]
    margem: List[float] = field(default_factory=list)    # [m0..mN] margem EBIT
    dividendos: List[float] = field(default_factory=list)  # [0, d1..dN] (legado, não usado)
    ev_ebit_saida: float = 8.0   # cross-check (EV/EBIT de saída implícito), não dirige o modelo
    net_debt_saida: float = 0.0
    taxa_desconto: float = 0.15  # WACC
    tax: float = 0.34            # alíquota p/ NOPAT = EBIT·(1−tax)
    # ── Drivers do FCFF (editáveis) ────────────────────────────────────────────
    da_pct: float = 0.04         # D&A como % da receita
    capex_pct: float = 0.05      # Capex como % da receita
    dso: float = 45.0            # dias de recebíveis (Receber/Receita·365)
    dio: float = 60.0            # dias de estoque (Estoque/CMV·365)
    dpo: float = 40.0            # dias de fornecedores (Fornecedores/CMV·365)
    cogs_pct: float = 0.60       # CMV como % da receita (base do giro)
    g_perp: float = 0.05         # crescimento na perpetuidade (Gordon, nominal BRL)


# ─── Engine ───────────────────────────────────────────────────────────────────

def _irr(flows: List[float], guess: float = 0.12) -> Optional[float]:
    """TIR via Newton-Raphson com fallback de bisseção. None se não convergir."""
    # Teto de taxa: TIRs acima disso não têm sentido econômico e só servem para
    # estourar (1+r)**i. Mantê-lo evita OverflowError no Newton divergente —
    # uma única empresa com fluxos degenerados não pode derrubar todo o valuation.
    _R_MAX = 1e6

    def npv(r: float) -> float:
        # (1+r)**i estoura para r enorme; trata como não-finito → empurra p/ bisseção
        try:
            return sum(cf / (1 + r) ** i for i, cf in enumerate(flows))
        except OverflowError:
            return float("inf")

    def dnpv(r: float) -> float:
        try:
            return sum(-i * cf / (1 + r) ** (i + 1) for i, cf in enumerate(flows))
        except OverflowError:
            return float("inf")

    r = guess
    for _ in range(100):
        f = npv(r)
        if abs(f) < 1e-9:
            return r
        d = dnpv(r)
        if d == 0 or not math.isfinite(f) or not math.isfinite(d):
            break
        r_new = r - f / d
        if r_new <= -0.9999:
            r_new = (r - 0.9999) / 2
        if r_new > _R_MAX:
            break  # Newton disparou → cai para a bisseção limitada
        if abs(r_new - r) < 1e-12:
            return r_new
        r = r_new

    lo, hi = -0.99, 10.0
    f_lo = npv(lo)
    if f_lo == 0:
        return lo
    for _ in range(300):
        mid = (lo + hi) / 2
        f_mid = npv(mid)
        if abs(f_mid) < 1e-9:
            return mid
        if (f_lo < 0) != (f_mid < 0):
            hi = mid
        else:
            lo, f_lo = mid, f_mid
    return None


# Limites do reinvestimento e do ROIC (guarda-corpos do FCFF — ver _fcff_series).
_ROIC_MIN, _ROIC_MAX = 0.05, 0.50     # ROIC efetivo confinado a faixa econômica plausível
                                      # (teto 50% acomoda compounders asset-light; corta artefatos)
_RR_MIN, _RR_MAX     = 0.0, 0.85      # taxa de reinvestimento (cresc. nunca consome >85% do NOPAT)
_ROIC_FALLBACK       = 0.12           # ROIC quando o pipeline não fornece

# Guarda-corpos do FCFE (financeiras): ROE plausível p/ banco/seguradora/B3.
_ROE_MIN, _ROE_MAX = 0.05, 0.40       # ROE efetivo (teto 40% acomoda seguradoras/B3 asset-light)
_ROE_FALLBACK      = 0.15             # ROE quando o pipeline não fornece
_BR_MAX            = 0.95             # retenção máx. (g/ROE) — acima disso captaria capital externo


def _clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))


def calcular_valuation(a: Anchors, p: Premissas) -> Dict:
    """DCF de FIRMA tradicional (FCFF explícito + perpetuidade de Gordon) + TIR.

    Build-up explícito por ano (t=1..n):
      EBIT → NOPAT = EBIT·(1−t) → +D&A − Capex − ΔCapital de Giro = FCFF
      · D&A e Capex como % da receita; ΔWC por prazos (DSO/DIO/DPO).
    Valor terminal (perpetuidade de Gordon com reinvestimento NORMALIZADO):
      FCFF_term = NOPAT_n·(1+g)·(1 − g/ROIC);  TV = FCFF_term / (WACC − g)
      (a normalização garante que a firma reinveste só o necessário p/ crescer g
       no steady state — evita 'sobre-investir para sempre' do %Capex explícito.)
      EV = Σ FCFF_t/(1+WACC)^t + TV/(1+WACC)^n ;  Equity = EV − Dív.Líq.

    TIR é o retorno da FIRMA (paga o EV atual, recebe os FCFF, vende a perpetuidade
    no fim). Propriedade preservada: upside = 0 ⟺ TIR = WACC.
    """
    n   = p.anos
    tax = p.tax
    r   = p.taxa_desconto                                  # WACC

    # ── Projeção operacional ───────────────────────────────────────────────────
    rev = [a.receita_ltm]
    for i in range(1, n + 1):
        rev.append(rev[-1] * (1 + p.cresc[i]))
    ebit  = [rev[i] * p.margem[i] for i in range(n + 1)]
    nopat = [ebit[i] * (1 - tax) for i in range(n + 1)]
    cogs  = [rev[i] * p.cogs_pct for i in range(n + 1)]    # CMV (base do giro)
    da    = [rev[i] * p.da_pct   for i in range(n + 1)]
    capex = [rev[i] * p.capex_pct for i in range(n + 1)]

    # ── Capital de giro por prazos (DSO/DIO/DPO) → ΔWC ─────────────────────────
    wc = []
    for i in range(n + 1):
        receber = rev[i]  * p.dso / 365.0
        estoque = cogs[i] * p.dio / 365.0
        fornec  = cogs[i] * p.dpo / 365.0
        wc.append(receber + estoque - fornec)
    dwc = [0.0] + [wc[i] - wc[i - 1] for i in range(1, n + 1)]

    # ── FCFF explícito (t=1..n) ────────────────────────────────────────────────
    fcff = [0.0]
    for i in range(1, n + 1):
        fcff.append(nopat[i] + da[i] - capex[i] - dwc[i])

    # ── Perpetuidade de Gordon com reinvestimento normalizado ──────────────────
    roic_eff = _clamp(a.roic if (a.roic and a.roic > 0) else _ROIC_FALLBACK, _ROIC_MIN, _ROIC_MAX)
    g_perp   = min(p.g_perp, r - 0.005)                    # garante g < WACC
    rr_perp  = _clamp(g_perp / roic_eff, _RR_MIN, 0.90)    # reinvestimento steady-state
    fcff_term = nopat[n] * (1 + g_perp) * (1 - rr_perp)    # FCFF do 1º ano perpétuo
    tv = fcff_term / (r - g_perp) if r > g_perp else 0.0

    # ── EV / Equity / Preço justo ──────────────────────────────────────────────
    pv_fcff = sum(fcff[i] / (1 + r) ** i for i in range(1, n + 1))
    pv_tv   = tv / (1 + r) ** n
    ev_justo     = pv_fcff + pv_tv
    equity_justo = ev_justo - a.net_debt
    preco_justo  = equity_justo / a.n_acoes if a.n_acoes else None
    upside = (preco_justo / a.preco - 1) if (preco_justo and a.preco) else None

    # ── TIR da firma: −EV hoje, FCFF intermediários, FCFF_n + venda da perpetuidade
    ev_atual = a.preco * a.n_acoes + a.net_debt
    flows = [-ev_atual] + [fcff[i] for i in range(1, n)] + [fcff[n] + tv]
    tir = _irr(flows)

    ev_ebit_atual      = (ev_atual / ebit[0]) if ebit[0] else None
    ev_ebit_saida_impl = (tv / ebit[n]) if ebit[n] else None    # cross-check do TV
    preco_saida        = ((tv - p.net_debt_saida) / a.n_acoes) if a.n_acoes else None

    return {
        "rev": rev, "ebit": ebit, "nopat": nopat, "da": da, "capex": capex,
        "dwc": dwc, "wc": wc, "fcff": fcff, "flows": flows,
        "tv": tv, "fcff_term": fcff_term, "g_perp": g_perp, "rr_perp": rr_perp,
        "pv_fcff": pv_fcff, "pv_tv": pv_tv,
        "ev_justo": ev_justo, "equity_justo": equity_justo,
        "tir": tir, "preco_justo": preco_justo, "upside": upside,
        "preco_saida": preco_saida, "ev_ebit_atual": ev_ebit_atual,
        "ev_ebit_saida_impl": ev_ebit_saida_impl,
    }


# ─── Valuation FCFE (financeiras: bancos, seguradoras, B3/corretoras) ──────────

def calcular_valuation_fcfe(a: Anchors, p: Premissas) -> Dict:
    """DCF de EQUITY (FCFE) — modelo para financeiras, onde o FCFF não se aplica
    (dívida/depósito é matéria-prima, não financiamento; não há EBIT/Capex/giro).

    Free Cash Flow to Equity com reinvestimento amarrado ao ROE (à la Damodaran):
      FCFE_t = LucroLíq_t · (1 − g_t/ROE)   [retém g/ROE p/ crescer a base de capital]
      LucroLíq cresce a g (trajetória decai ao g perpétuo).
    Perpetuidade de Gordon:
      FCFE_term = LL_n·(1+g)·(1 − g/ROE);  TV = FCFE_term/(Re − g)
    Valor:
      Equity = Σ FCFE_t/(1+Re)^t + TV/(1+Re)^n ;  Preço justo = Equity/ações.

    Desconto ao CUSTO DE EQUITY (Re = Rf + β·ERP), NÃO ao WACC — para um banco a
    alavancagem é o próprio negócio. `p.taxa_desconto` carrega Re aqui.
    TIR é o retorno do acionista: paga o market cap hoje, recebe os FCFE + venda da
    perpetuidade. Propriedade preservada: upside = 0 ⟺ TIR = Re.
    """
    n  = p.anos
    re = p.taxa_desconto                                   # custo de equity (Re)

    roe_eff = _clamp(a.roe if (a.roe and a.roe > 0) else _ROE_FALLBACK, _ROE_MIN, _ROE_MAX)

    # ── Projeção do lucro líquido ──────────────────────────────────────────────
    ll = [a.lucro_liq_ltm]
    for i in range(1, n + 1):
        ll.append(ll[-1] * (1 + p.cresc[i]))

    # ── FCFE explícito: retém g/ROE p/ financiar o crescimento da base de capital
    fcfe = [0.0]
    br_ser = [0.0]
    for i in range(1, n + 1):
        br = _clamp(p.cresc[i] / roe_eff, 0.0, _BR_MAX)    # retenção (b = g/ROE)
        br_ser.append(br)
        fcfe.append(ll[i] * (1 - br))

    # ── Perpetuidade de Gordon (reinvestimento normalizado g/ROE) ──────────────
    g_perp  = min(p.g_perp, re - 0.005)                    # garante g < Re
    br_perp = _clamp(g_perp / roe_eff, 0.0, _BR_MAX)
    fcfe_term = ll[n] * (1 + g_perp) * (1 - br_perp)
    tv = fcfe_term / (re - g_perp) if re > g_perp else 0.0

    # ── Equity / Preço justo ───────────────────────────────────────────────────
    pv_fcfe = sum(fcfe[i] / (1 + re) ** i for i in range(1, n + 1))
    pv_tv   = tv / (1 + re) ** n
    equity_justo = pv_fcfe + pv_tv
    preco_justo  = equity_justo / a.n_acoes if a.n_acoes else None
    upside = (preco_justo / a.preco - 1) if (preco_justo and a.preco) else None

    # ── TIR do acionista: −market cap hoje, FCFE intermediários, FCFE_n + venda da perp.
    mkt_cap = a.preco * a.n_acoes
    flows = [-mkt_cap] + [fcfe[i] for i in range(1, n)] + [fcfe[n] + tv]
    tir = _irr(flows)

    pl_justo_pvp = (equity_justo / a.pl) if a.pl else None  # P/VP implícito do valor justo
    pvp_atual    = (mkt_cap / a.pl) if a.pl else None

    return {
        "modelo": "FCFE", "ll": ll, "fcfe": fcfe, "br_ser": br_ser, "flows": flows,
        "roe_eff": roe_eff, "tv": tv, "fcfe_term": fcfe_term,
        "g_perp": g_perp, "br_perp": br_perp,
        "pv_fcfe": pv_fcfe, "pv_tv": pv_tv,
        "equity_justo": equity_justo, "preco_justo": preco_justo, "upside": upside,
        "tir": tir, "pvp_atual": pvp_atual, "pvp_justo": pl_justo_pvp,
    }


def premissas_default_fcfe(a: Anchors, anos: int = 5) -> Premissas:
    """Premissas do FCFE: crescimento do lucro decai do histórico (ou implícito por
    ROE×retenção) até o g perpétuo. `taxa_desconto` é preenchido depois com Re."""
    g_perp_val = 0.05
    # Crescimento inicial: usa CAGR histórico do lucro/receita; fallback = ROE·retenção
    # sustentável (g = ROE·(1−payout); sem payout, assume retenção 0.5).
    if a.cagr_hist is not None:
        g0 = a.cagr_hist
    elif a.roe:
        g0 = a.roe * 0.5
    else:
        g0 = 0.10
    g0 = max(0.0, min(g0, 0.30))
    cresc = [0.0]
    g = g0
    for _ in range(anos):
        cresc.append(round(g, 4))
        g = max(g_perp_val, g * 0.85)
    margem = [round(a.margem_ebit, 4)] * (anos + 1)        # não usado no FCFE, mantém shape
    return Premissas(
        anos=anos, base_year=date.today().year,
        cresc=cresc, margem=margem, dividendos=[0.0] * (anos + 1),
        taxa_desconto=0.15, tax=0.34, g_perp=g_perp_val,
    )


# ─── Custo de capital (WACC) ──────────────────────────────────────────────────

@dataclass
class WACC:
    """Resultado do cálculo de custo médio ponderado de capital (pós-imposto)."""
    beta: float                  # beta usado (ajustado, por padrão)
    rf: float                    # taxa livre de risco
    erp: float                   # prêmio de risco de mercado (equity risk premium)
    custo_equity: float          # Re = Rf + β·ERP (CAPM)
    custo_divida: float          # Rd pré-imposto
    tax: float                   # alíquota efetiva (escudo fiscal da dívida)
    peso_equity: float           # E/(E+D)
    peso_divida: float           # D/(E+D)
    wacc: float                  # WACC pós-imposto


def calcular_beta(datas_a: List[str], close_a: List[float],
                  datas_b: List[str], close_b: List[float]):
    """Beta via regressão de retornos diários da ação vs benchmark (IBOV).

    β = Cov(r_ação, r_mercado) / Var(r_mercado).
    Retorna (beta_raw, beta_ajustado, n_obs). O ajustado usa a correção de
    Blume (0.67·β + 0.33), que puxa para a média de mercado (1.0) e é prática
    padrão em valuation. Retorna (None, None, 0) se houver dados insuficientes.
    """
    mb = dict(zip(datas_b, close_b))
    pares = [(close_a[i], mb[d]) for i, d in enumerate(datas_a) if d in mb]
    if len(pares) < 30:
        return None, None, 0

    ra, rb = [], []
    for i in range(1, len(pares)):
        pa0, pb0 = pares[i - 1]
        pa1, pb1 = pares[i]
        if pa0 and pb0 and pa0 > 0 and pb0 > 0:
            ra.append(pa1 / pa0 - 1)
            rb.append(pb1 / pb0 - 1)

    n = len(ra)
    if n < 30:
        return None, None, 0

    mean_a = sum(ra) / n
    mean_b = sum(rb) / n
    cov = sum((ra[i] - mean_a) * (rb[i] - mean_b) for i in range(n)) / n
    var = sum((rb[i] - mean_b) ** 2 for i in range(n)) / n
    if var == 0:
        return None, None, 0

    beta = cov / var
    beta_adj = 0.67 * beta + 0.33          # ajuste de Blume
    return round(beta, 3), round(beta_adj, 3), n


def calcular_wacc(mkt_cap: float, divida_bruta: float, beta: float,
                  rf: float = 0.105, erp: float = 0.075,
                  custo_divida: float = 0.12, tax: float = 0.34) -> WACC:
    """WACC pós-imposto. `mkt_cap` e `divida_bruta` na mesma unidade (R$ mi).

    Custo do equity via CAPM: Re = Rf + β·ERP.
    Custo da dívida líquido de imposto: Rd·(1−tax).
    Pesos pelo valor de mercado do equity (E) e dívida bruta contábil (D).
    Defaults Brasil: Rf 10.5% · ERP 7.5% · Rd 12% · imposto 34% (IRPJ+CSLL).
    """
    re = rf + beta * erp
    e = max(mkt_cap, 0.0)
    d = max(divida_bruta, 0.0)
    v = e + d
    if v <= 0:
        we, wd = 1.0, 0.0
    else:
        we, wd = e / v, d / v
    wacc = we * re + wd * custo_divida * (1 - tax)
    return WACC(beta=beta, rf=rf, erp=erp, custo_equity=re,
                custo_divida=custo_divida, tax=tax,
                peso_equity=we, peso_divida=wd, wacc=wacc)


# ─── Carregamento automático de âncoras ───────────────────────────────────────

def montar_anchors(ticker: str, anos_dfp: int = 7) -> Anchors:
    """Preenche as âncoras a partir do pipeline (CVM + FRE + Fundamentus)."""
    import csv
    from cvm_client import load_companies_bulk
    from indicadores_empresas import calcular_indicadores
    from acoes_circulacao import load_shares
    from fundamentus_client import load_fundamentus

    ticker = ticker.strip().upper()
    with open("empresas_lista.csv", encoding="utf-8", newline="") as f:
        rows = [e for e in csv.DictReader(f)
                if (e.get("ticker_b3") or "").strip().upper() == ticker]
    if not rows:
        raise SystemExit(f"Ticker {ticker} não encontrado em empresas_lista.csv")
    row = rows[0]
    cnpj = "".join(c for c in (row.get("cnpj") or "") if c.isdigit())

    emp = list(load_companies_bulk([row], anos_dfp=anos_dfp).values())[0]
    res = calcular_indicadores(emp)
    cb = res.get("campos_brutos") or {}
    ind = res.get("ind") or {}

    MI = 1e6
    receita = (cb.get("receita") or 0.0) / MI
    ebit = (cb.get("ebit") or 0.0) / MI
    net_debt = ((cb.get("divida_cp") or 0.0) + (cb.get("divida_lp") or 0.0)
                - (cb.get("caixa") or 0.0)) / MI
    lucro = (cb.get("lucro_liq") or 0.0) / MI

    # Ações em circulação (FRE, mais recente)
    sl = load_shares(list(range(2019, date.today().year + 1)))
    n_acoes_raw = sl.asof(cnpj, date.today().isoformat())
    n_acoes = (n_acoes_raw / MI) if n_acoes_raw else 0.0

    # Preço e comparáveis (Fundamentus)
    fund = load_fundamentus().get(ticker, {})
    preco = fund.get("cotacao") or 0.0
    roe = (fund.get("roe") / 100.0) if fund.get("roe") is not None else None

    # CAGR histórico de receita (até 3 anos)
    hist = (res.get("historico_brutos") or {}).get("receita") or []
    hist = [h for h in hist if h]
    cagr = None
    if len(hist) >= 2 and hist[0] > 0:
        cagr = (hist[-1] / hist[0]) ** (1 / (len(hist) - 1)) - 1

    margem = (ebit / receita) if receita else 0.0
    mktcap = preco * n_acoes
    ev_ebit = ((mktcap + net_debt) / ebit) if ebit else None
    pe = (mktcap / lucro) if lucro else None
    roic = (ind.get("roic") / 100.0) if ind.get("roic") is not None else None
    da_ltm    = abs(cb.get("da") or 0.0) / MI
    capex_ltm = abs(cb.get("capex") or 0.0) / MI
    cogs_ltm  = abs(cb.get("custo_vendas") or 0.0) / MI
    receber   = (cb.get("contas_receber") or 0.0) / MI
    estoque   = (cb.get("estoques") or 0.0) / MI
    fornec    = (cb.get("fornecedores") or 0.0) / MI

    return Anchors(
        ticker=ticker, nome=res.get("nome") or emp.nome,
        preco=preco, n_acoes=n_acoes, net_debt=net_debt,
        receita_ltm=receita, ebit_ltm=ebit, margem_ebit=margem,
        ev_ebit_atual=ev_ebit, pe_atual=pe, roe=roe, cagr_hist=cagr, roic=roic,
        da_ltm=da_ltm, capex_ltm=capex_ltm, cogs_ltm=cogs_ltm,
        receber=receber, estoque=estoque, fornecedores=fornec,
        fontes={
            "preco": "Fundamentus (cotação)",
            "n_acoes": "CVM FRE (capital integralizado)",
            "net_debt": f"CVM LTM ({res.get('data_ref','')})",
            "receita_ltm": f"CVM LTM ({res.get('data_ref','')})",
            "ebit_ltm": f"CVM LTM ({res.get('data_ref','')})",
        },
    )


def premissas_default(a: Anchors, anos: int = 5) -> Premissas:
    """Defaults editáveis derivados do LTM: crescimento decai do histórico até o g
    perpétuo (handoff suave p/ a perpetuidade), margem atual, D&A/Capex/CMV como %
    da receita e giro por prazos (DSO/DIO/DPO). Horizonte explícito de 5 anos."""
    g_perp_val = 0.05
    g0 = a.cagr_hist if a.cagr_hist is not None else 0.10
    g0 = max(0.0, min(g0, 0.30))                 # limita 0–30%
    cresc = [0.0]
    g = g0
    for _ in range(anos):
        cresc.append(round(g, 4))
        g = max(g_perp_val, g * 0.85)            # decai 15% a.a., piso = g perpétuo
    margem = [round(a.margem_ebit, 4)] * (anos + 1)

    rev  = a.receita_ltm or 0.0
    cogs = a.cogs_ltm or (rev * 0.60)
    da_pct   = _clamp(a.da_ltm / rev, 0.0, 0.30) if rev else 0.04
    capex_pct = _clamp(a.capex_ltm / rev, 0.0, 0.40) if rev else 0.05
    cogs_pct = _clamp(cogs / rev, 0.05, 0.95) if rev else 0.60
    dso = _clamp(a.receber / rev * 365.0, 0.0, 365.0) if rev else 45.0
    dio = _clamp(a.estoque / cogs * 365.0, 0.0, 730.0) if cogs else 0.0
    dpo = _clamp(a.fornecedores / cogs * 365.0, 0.0, 365.0) if cogs else 0.0
    ev_saida = round(a.ev_ebit_atual, 1) if a.ev_ebit_atual and a.ev_ebit_atual > 0 else 8.0

    return Premissas(
        anos=anos, base_year=date.today().year,
        cresc=cresc, margem=margem, dividendos=[0.0] * (anos + 1),
        ev_ebit_saida=ev_saida, net_debt_saida=round(a.net_debt, 1),
        taxa_desconto=0.15, tax=0.34,
        da_pct=round(da_pct, 4), capex_pct=round(capex_pct, 4), cogs_pct=round(cogs_pct, 4),
        dso=round(dso, 0), dio=round(dio, 0), dpo=round(dpo, 0), g_perp=g_perp_val,
    )


# ─── Geração do Excel (formato playbook) ──────────────────────────────────────

def gerar_excel(a: Anchors, p: Premissas, path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    AZUL = Font(name="Arial", color="0000FF")      # inputs / premissas (editáveis)
    PRETO = Font(name="Arial", color="000000")     # fórmulas
    BOLD = Font(name="Arial", bold=True)
    HEADER = Font(name="Arial", bold=True, color="FFFFFF")
    AMARELO = PatternFill("solid", fgColor="FFF2CC")
    VERDE = PatternFill("solid", fgColor="284B23")
    fmt_rs = '#,##0;(#,##0);"-"'
    fmt_mult = '0.0"x"'
    fmt_pct = '0.0%'
    fmt_preco = 'R$ #,##0.00'

    n = p.anos
    c0 = 4                                          # coluna D = ano base
    cT = c0 + n                                     # coluna terminal
    L = get_column_letter
    base = L(c0)
    term = L(cT)

    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation"

    def put(cell, val, font=PRETO, fmt=None, fill=None, comment=None, align=None):
        c = ws[cell]
        c.value = val
        c.font = font
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        if align: c.alignment = Alignment(horizontal=align)
        if comment: c.comment = Comment(comment, "Zelen")

    # Cabeçalho
    put("C2", f"Valuation — {a.ticker}", BOLD)
    for i in range(n + 1):
        # Ano como número com formato '0' (evita "2.025" e mantém COUNT/numérico)
        put(f"{L(c0+i)}2", p.base_year + i, BOLD, fmt="0", align="center")

    col1 = L(c0 + 1)                                    # 1ª coluna de projeção (ano 1)

    # Preço / ações / dívida
    put("C3", "Preço (R$)")
    put(f"{base}3", round(a.preco, 2), AZUL, fmt_preco, comment=a.fontes.get("preco"))
    put(f"{term}3", f"=($D$35-{base}5)/{base}4", PRETO, fmt_preco,
        comment="Preço de saída ≈ (Valor Terminal − Dív.Líq) / ações")
    put("C4", "n* ações (mi)")
    put(f"{base}4", round(a.n_acoes, 1), AZUL, '#,##0.0', comment=a.fontes.get("n_acoes"))
    put("C5", "Net Debt (R$ mi)")
    put(f"{base}5", round(a.net_debt, 0), AZUL, fmt_rs, comment=a.fontes.get("net_debt"))

    # Receita (base âncora; projeções por fórmula)
    put("C6", "Receita (R$ mi)")
    put(f"{base}6", round(a.receita_ltm, 0), AZUL, fmt_rs, comment=a.fontes.get("receita_ltm"))
    for i in range(1, n + 1):
        put(f"{L(c0+i)}6", f"={L(c0+i-1)}6*(1+{L(c0+i)}7)", PRETO, fmt_rs)
    # Crescimento (premissa azul, por ano)
    put("C7", "Cresc. (%)")
    for i in range(1, n + 1):
        put(f"{L(c0+i)}7", p.cresc[i], AZUL, fmt_pct, fill=AMARELO)
    # Margem EBIT (premissa azul, por ano)
    put("C8", "Mrg. EBIT (%)")
    for i in range(n + 1):
        put(f"{L(c0+i)}8", p.margem[i], AZUL, fmt_pct, fill=AMARELO)
    # EBIT = receita × margem
    put("C9", "EBIT (R$ mi)")
    for i in range(n + 1):
        put(f"{L(c0+i)}9", f"={L(c0+i)}6*{L(c0+i)}8", PRETO, fmt_rs)
    # NOPAT = EBIT × (1 − imposto)
    put("C10", "NOPAT (R$ mi)")
    for i in range(n + 1):
        put(f"{L(c0+i)}10", f"={L(c0+i)}9*(1-$D$25)", PRETO, fmt_rs)
    # CMV (base do capital de giro)
    put("C11", "CMV (R$ mi)")
    for i in range(n + 1):
        put(f"{L(c0+i)}11", f"={L(c0+i)}6*$D$28", PRETO, fmt_rs)
    # (+) D&A = receita × D&A%
    put("C12", "(+) D&A")
    for i in range(n + 1):
        put(f"{L(c0+i)}12", f"={L(c0+i)}6*$D$26", PRETO, fmt_rs)
    # (−) Capex = receita × Capex%
    put("C13", "(−) Capex")
    for i in range(n + 1):
        put(f"{L(c0+i)}13", f"={L(c0+i)}6*$D$27", PRETO, fmt_rs)
    # Capital de giro por prazos (Receber + Estoque − Fornecedores)
    put("C14", "Cap. Giro (R$ mi)")
    for i in range(n + 1):
        put(f"{L(c0+i)}14",
            f"={L(c0+i)}6*$D$29/365+{L(c0+i)}11*$D$30/365-{L(c0+i)}11*$D$31/365", PRETO, fmt_rs)
    # (−) Δ Capital de Giro
    put("C15", "(−) Δ Cap. Giro")
    for i in range(1, n + 1):
        put(f"{L(c0+i)}15", f"={L(c0+i)}14-{L(c0+i-1)}14", PRETO, fmt_rs)
    # (=) FCFF = NOPAT + D&A − Capex − ΔGiro
    put("C16", "(=) FCFF (R$ mi)")
    for i in range(1, n + 1):
        put(f"{L(c0+i)}16", f"={L(c0+i)}10+{L(c0+i)}12-{L(c0+i)}13-{L(c0+i)}15", PRETO, fmt_rs)
    # Fluxos da firma: entrada = −EV atual; intermediários = FCFF; saída = FCFF_n + TV
    put("C17", "Fluxos (R$ mi)")
    put(f"{base}17", f"=-({base}3*{base}4+{base}5)", PRETO, fmt_rs)
    for i in range(1, n):
        put(f"{L(c0+i)}17", f"={L(c0+i)}16", PRETO, fmt_rs)
    put(f"{term}17", f"={term}16+$D$35", PRETO, fmt_rs)   # FCFF_n + Valor Terminal

    # ── Resultados ────────────────────────────────────────────────────────────
    put("C19", "TIR (firma)", BOLD)
    put(f"{base}19", f"=IRR({base}17:{term}17)", BOLD, fmt_pct)
    put("C20", "EV Justo (R$ mi)", BOLD)
    # NPV dos FCFF (anos 1..n); o fluxo terminal já inclui a perpetuidade (TV)
    put(f"{base}20", f"=NPV($D$34,{col1}17:{term}17)", BOLD, fmt_rs)
    put("C21", "Equity Justo (R$ mi)", BOLD)
    put(f"{base}21", f"={base}20-{base}5", BOLD, fmt_rs)
    put("C22", "Preço Justo (R$)", BOLD)
    put(f"{base}22", f"={base}21/{base}4", BOLD, fmt_preco)
    put("C23", "Upside", BOLD)
    put(f"{base}23", f"={base}22/{base}3-1", BOLD, fmt_pct)
    ws[f"{base}23"].fill = AMARELO

    # ── Premissas-chave (azul = editável) ─────────────────────────────────────
    fmt_dias = '#,##0"d"'
    roic_def = round(a.roic, 4) if a.roic else 0.12
    prem_rows = [
        (25, "Imposto (%)",    round(p.tax, 4),          fmt_pct,  "Alíquota p/ NOPAT = EBIT·(1−imposto)"),
        (26, "D&A (% rec.)",   round(p.da_pct, 4),       fmt_pct,  "Depreciação & Amortização como % da receita"),
        (27, "Capex (% rec.)", round(p.capex_pct, 4),    fmt_pct,  "Capex como % da receita"),
        (28, "CMV (% rec.)",   round(p.cogs_pct, 4),     fmt_pct,  "Custo dos produtos vendidos — base do giro"),
        (29, "Receber (DSO)",  round(p.dso, 0),          fmt_dias, "Dias de recebíveis (sobre receita)"),
        (30, "Estoque (DIO)",  round(p.dio, 0),          fmt_dias, "Dias de estoque (sobre CMV)"),
        (31, "Fornec. (DPO)",  round(p.dpo, 0),          fmt_dias, "Dias de fornecedores (sobre CMV)"),
        (32, "ROIC (%)",       roic_def,                 fmt_pct,  "ROIC p/ reinvestimento na perpetuidade"),
        (33, "g perpétuo (%)", round(p.g_perp, 4),       fmt_pct,  "Crescimento na perpetuidade (Gordon)"),
        (34, "WACC (%)",       round(p.taxa_desconto, 4), fmt_pct, "Custo de capital — taxa de desconto"),
    ]
    for r, lbl, val, fmt, cm in prem_rows:
        put(f"C{r}", lbl)
        put(f"{base}{r}", val, AZUL, fmt, fill=AMARELO, comment=cm)
    # Valor Terminal (perpetuidade de Gordon, reinvestimento normalizado g/ROIC)
    put("C35", "Valor Terminal (TV)", BOLD)
    put(f"{base}35", f"=({term}10*(1+$D$33)*(1-$D$33/$D$32))/($D$34-$D$33)", BOLD, fmt_rs,
        comment="Perpetuidade: FCFF_term/(WACC−g), FCFF_term = NOPAT_n·(1+g)·(1−g/ROIC)")

    # ── Análise Setorial (comparáveis) ───────────────────────────────────────
    put("C37", "Análise Setorial", BOLD)
    for col, lbl in zip(("D", "E", "F", "G"), ("P/E", "EV/EBIT", "ROE", "Cresc.")):
        put(f"{col}37", lbl, BOLD, align="center")
    put("C38", "Empresa analisada")
    put("D38", round(a.pe_atual, 1) if a.pe_atual else None, PRETO, fmt_mult)
    put("E38", round(a.ev_ebit_atual, 1) if a.ev_ebit_atual else None, PRETO, fmt_mult)
    put("F38", round(a.roe, 4) if a.roe is not None else None, PRETO, fmt_pct)
    put("G38", round(a.cagr_hist, 4) if a.cagr_hist is not None else None, PRETO, fmt_pct)
    put("C39", "Par 1 (preencher)")
    put("C40", "Par 2 (preencher)")
    for r in (39, 40):
        put(f"D{r}", None, AZUL, fmt_mult, fill=AMARELO)
        put(f"E{r}", None, AZUL, fmt_mult, fill=AMARELO)
        put(f"F{r}", None, AZUL, fmt_pct, fill=AMARELO)
        put(f"G{r}", None, AZUL, fmt_pct, fill=AMARELO)
    put("C41", "Média", BOLD)
    for col, fmt in zip(("D", "E", "F", "G"), (fmt_mult, fmt_mult, fmt_pct, fmt_pct)):
        put(f"{col}41", f"=IFERROR(AVERAGE({col}39:{col}40),\"\")", PRETO, fmt)
    put("C42", "Diferença", BOLD)
    for col, fmt in zip(("D", "E", "F", "G"), (fmt_pct, fmt_pct, fmt_pct, fmt_pct)):
        put(f"{col}42", f"=IFERROR({col}38/{col}41-1,\"\")", PRETO, fmt_pct)

    # Estética
    ws.column_dimensions["C"].width = 20
    for i in range(n + 1):
        ws.column_dimensions[L(c0 + i)].width = 11
    # Legenda
    put("C44", "Azul = premissa editável · Preto = fórmula · Amarelo = premissa-chave",
        Font(name="Arial", italic=True, size=9, color="808080"))
    put("C45", "Modelo: DCF de firma — FCFF = NOPAT + D&A − Capex − ΔGiro, VP ao WACC + perpetuidade de Gordon.",
        Font(name="Arial", italic=True, size=9, color="808080"))
    put("C46", f"Âncoras automáticas: {a.nome} · gerado {date.today().isoformat()}",
        Font(name="Arial", italic=True, size=9, color="808080"))

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ─── CLI ──────────────────────────────────────────────────────────────────────

def main() -> int:
    import io
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8",
                                      errors="replace", line_buffering=True)
    ap = argparse.ArgumentParser(description="Valuation semi-automatizado — Zelen")
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--anos", type=int, default=5)
    ap.add_argument("--saida-mult", type=float, default=None,
                    help="EV/EBIT de saída (default = EV/EBIT atual)")
    ap.add_argument("--desconto", type=float, default=0.15)
    ap.add_argument("--output", type=Path, default=None)
    args = ap.parse_args()

    print(f"\nCarregando âncoras de {args.ticker.upper()}…", file=sys.stderr)
    a = montar_anchors(args.ticker)
    p = premissas_default(a, anos=args.anos)
    if args.saida_mult is not None:
        p.ev_ebit_saida = args.saida_mult
    p.taxa_desconto = args.desconto

    r = calcular_valuation(a, p)

    def pct(x): return f"{x*100:.1f}%" if x is not None else "—"
    print(f"\n  {a.ticker} — {a.nome}")
    print(f"  {'─'*46}")
    print(f"  Preço atual ........ R$ {a.preco:.2f}")
    print(f"  Ações .............. {a.n_acoes:,.1f} mi")
    print(f"  Net Debt ........... R$ {a.net_debt:,.0f} mi")
    print(f"  Receita LTM ........ R$ {a.receita_ltm:,.0f} mi")
    print(f"  EBIT LTM ........... R$ {a.ebit_ltm:,.0f} mi  (margem {pct(a.margem_ebit)})")
    print(f"  EV/EBIT atual ...... {r['ev_ebit_atual']:.1f}x")
    print(f"  {'─'*46}")
    print(f"  Premissas: cresc {[pct(g) for g in p.cresc[1:]]}, "
          f"saída {p.ev_ebit_saida:.1f}x, desconto {pct(p.taxa_desconto)}")
    print(f"  {'─'*46}")
    print(f"  TIR ................ {pct(r['tir'])}")
    print(f"  Preço Justo ........ R$ {r['preco_justo']:.2f}" if r['preco_justo'] else "  Preço Justo: —")
    print(f"  Upside ............. {pct(r['upside'])}")
    print(f"  Preço saída ({p.base_year+p.anos}) . R$ {r['preco_saida']:.2f}")

    out = args.output or (Path("relatorios") / f"valuation_{a.ticker}.xlsx")
    path = gerar_excel(a, p, out)
    print(f"\n  Excel: {path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
